from openpyxl import Workbook


class ExcelUtil:
    def __init__(self):
        self.wb = Workbook()

    def create_sheet(self, sheet_name):
        self.wb.create_sheet(sheet_name)

    def insert_data(self, sheet_name, row, col, volumn):
        sheet = self.wb[sheet_name]
        sheet.cell(row=row, column=col).value = volumn

    def insert_line_date(self, sheet_name, row, volums):
        sheet = self.wb[sheet_name]
        for index, volumn in enumerate(volums):
            sheet.cell(row=row, column=index + 1).value = volumn

    def save(self, path):
        self.wb.save(path)
